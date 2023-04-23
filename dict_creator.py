'''
This script is developed with python 3.6 and it reads a downloaded file and uploads the data to s3 as a json file.

AWS credentials should in place and write access for the configured bucket should be provided in order to run this script
'''
import sys
import json
import requests
import time
import xlrd
import boto3
from openpyxl import Workbook, load_workbook

# Initializing s3 resource
s3 = boto3.resource('s3')

# get information from config.json file
with open('config.json', 'r') as config_file:
    data = json.load(config_file)

bucket_name = data['bucket_name']
xlsx_url = data['url']
split_url = xlsx_url.split('/')
file_name = split_url[len(split_url)-1]  # Getting file name to save file later
sheet_name = data['sheet_name']          # sheet to read
return_url = data['return_url'] if data['retun_url_bool'] else None # Initializing return url based on retun_url_bool
response = requests.get(xlsx_url) 

# Stops executing the function if file download url returns anything other than 200 (status_code)
if not response.ok:
    sys.exit('Service Unavailable')

# checks for 'xls' file format
if file_name[-3:] == 'xls':      
    # Converting xls to xlsx format
    xlsBook = xlrd.open_workbook(file_contents=response.content)
    workbook = Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name
    
        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col) if xlsSheet.cell_value(row, col) != '' else None

    # Saves data file in xlsx format 
    workbook.save(file_name[:-3]+'xlsx')

#Checks for xlsx file format
elif file_name[-4:] == 'xlsx':
    with open(file_name, 'w') as f: # Saving and loading the data file
        f.write(response.content)
    workbook = load_workbook(file_name)

# Stops executing if the file is not in xls or xlsx format  
else:
    sys.exit('Unsupported file format')

try:
    sheet = workbook[sheet_name] # loads 'MICs List by CC' sheet mentioned in config.json
except KeyError:
    sys.exit('Specified sheet is missing') # Stops executing if 'MICs List by CC' sheet is missing

data_list = []

# list of all the column names
key_values = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column+1)]  

# Generating dictionary of all the rows with column name as keys
for i in range(2, sheet.max_row+1):
    row_dict = {key_values[x-1]: sheet.cell(row=i, column=x).value for x in range(1, sheet.max_column+1)}
    data_list.append(row_dict)


# Saving json data on s3
try:
    time_stamp_string = time.strftime("%Y%m%d-%H%M%S")  # To generate unique filename
    s3.Bucket(bucket_name).put_object(Key='data_' + time_stamp_string +'.json', Body=json.dumps(data_list, indent=4))
    if return_url:
        msg = return_url.format('data_' + time_stamp_string +'.json')
    else:
        msg = "File {} uploaded".format('data_' + time_stamp_string +'.json')
except Exception as e:
    print (e)
    msg = 'Upload failed, please check your S3 configuration'

print (msg) # final status of the script
